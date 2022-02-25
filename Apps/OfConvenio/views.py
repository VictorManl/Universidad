from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView, TemplateView, DetailView
from django.http.response import HttpResponse, HttpResponseRedirect
from .models import Convenios, ProyectosProyeccionSocial
from .forms import listaFormulario, formC
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


# Create your views here.

# Login

def Login(request):
    return render(request, 'Login.html')


class index(TemplateView):
    page_title = 'Inicio'
    template_name = 'index.html'


    def get_context_data(self, **kwargs):
        context = super(index, self).get_context_data(**kwargs)
        context['page_title'] = self.page_title
        return context


class InicioConvenio(ListView):
    model = Convenios
    template_name = 'Convenio/InicioConvenio.html'
    profile_list = Convenios.objects.get_queryset().order_by('codigo')
    paginate_by = 5


class crearC(CreateView):
    model = Convenios
    form_class = listaFormulario
    page_title = 'Nuevo Convenio'
    template_name = 'Convenio/crearConvenio.html'
    success_url = reverse_lazy('inicioConvenio')

    def get_context_data(self, **kwargs):
        context = super(crearC, self).get_context_data(**kwargs)
        context['page_title'] = self.page_title
        return context


class editarC(UpdateView):
    model = Convenios
    form_class = listaFormulario
    page_title = 'Editar Convenio'
    success_url = reverse_lazy('inicioConvenio')
    context_object_name = 'obj'
    template_name = 'Convenio/convenio.html'

    def get_context_data(self, **kwargs):
        context = super(editarC, self).get_context_data(**kwargs)
        context['page_title'] = self.page_title
        return context


class reporteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        # Crear el título en la hoja
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
        ws['B1'].font = Font(name='Century Gothic', size=16, bold=True)
        ws['B1'] = 'Reporte de Convenios'

        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:I1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 35
        ws.column_dimensions['I'].width = 50

        # Crear la cabecera
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['B2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['B2'] = 'Fecha'

        ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['C2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['C2'] = 'Numero de Cc'

        ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['D2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['D2'] = 'Numero del convenio'

        ws['E2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['E2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['E2'] = 'Nombre del estudiante'

        ws['F2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['F2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['F2'] = 'Telefono'

        ws['G2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['G2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['G2'] = 'Municipio'

        ws['H2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['H2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['H2'] = 'Programa'

        ws['I2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['I2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['I2'] = 'Facultad'

        contador = 3
        for convenios in Convenios.objects.all():
            # Pintamos los datos en el reporte
            ws.cell(row=contador, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=2).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=2).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=2).value = convenios.fecha

            ws.cell(row=contador, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=3).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=3).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=3).value = convenios.documento

            ws.cell(row=contador, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=4).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=4).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=4).value = convenios.convenio

            ws.cell(row=contador, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=5).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=5).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=5).value = convenios.nombre

            ws.cell(row=contador, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=6).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=6).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=6).value = convenios.telefono

            ws.cell(row=contador, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=7).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=7).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=7).value = convenios.municipio

            ws.cell(row=contador, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=8).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=8).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=8).value = convenios.programa

            ws.cell(row=contador, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=9).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=9).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=9).value = convenios.facultad
            contador += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "ReporteConvenios.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


# Proyeccion

class proyeccionInicio(ListView):
    model = ProyectosProyeccionSocial
    template_name = 'Proyeccion/inicioProyeccion.html'
    profile_list = ProyectosProyeccionSocial.objects.get_queryset().order_by('codigo')
    paginate_by = 5


class CrearProyecto(CreateView):
    model = ProyectosProyeccionSocial
    form_class = formC
    page_title = 'Nuevo Proyecto'
    template_name = "Proyeccion/crearProyecto.html"
    success_url = reverse_lazy('inicioProyeccion')

    def get_context_data(self, **kwargs):
        context = super(CrearProyecto, self).get_context_data(**kwargs)
        context['page_title'] = self.page_title
        return context


class editarProyecto(UpdateView):
    model = ProyectosProyeccionSocial
    form_class = formC
    page_title = 'Editar Proyecto'
    success_url = reverse_lazy('inicioProyeccion')
    context_object_name = 'obj'
    template_name = 'Proyeccion/crearProyecto.html'

    def get_context_data(self, **kwargs):
        context = super(editarProyecto, self).get_context_data(**kwargs)
        context['page_title'] = self.page_title


class verProyecto(DetailView):
    model = ProyectosProyeccionSocial
    page_title = 'Ver Proyecto'
    template_name = 'Proyeccion/ListaProyecto.html'

    def get_context_data(self, **kwargs):
        context = super(verProyecto, self).get_context_data(**kwargs)
        context['page_title'] = self.page_title
        return context


class proyeccionReporte(TemplateView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        # Crear el título en la hoja
        ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
        ws['B1'].font = Font(name='Century Gothic', size=16, bold=True)
        ws['B1'] = 'Reporte de Proyectos'
        # Cambiar caracteristicas de las celdas
        ws.merge_cells('B1:M1')

        ws.row_dimensions[1].height = 25

        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 55
        ws.column_dimensions['J'].width = 35
        ws.column_dimensions['K'].width = 35
        ws.column_dimensions['L'].width = 35
        ws.column_dimensions['M'].width = 150

        # Crear la cabecera
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['B2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['B2'] = 'Fecha'

        ws['C2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['C2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['C2'] = 'Programa'

        ws['D2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['D2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['D2'] = 'Ciudad'

        ws['E2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['E2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['E2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['E2'] = 'Institucion'

        ws['F2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['F2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['F2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['F2'] = 'Numero de estudiantes'

        ws['G2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['G2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['G2'] = 'Tipo de comunidad'

        ws['H2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['H2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['H2'] = 'Organizacion'

        ws['I2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['I2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['I2'] = 'Facultad'

        ws['J2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['J2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['J2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['J2'] = 'Nombre del proyecto'

        ws['K2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['K2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['K2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['K2'] = 'Nombre del docente'

        ws['L2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['L2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['L2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['L2'] = 'Codigo BPUT'

        ws['M2'].alignment = Alignment(horizontal="justify", vertical="justify")
        ws['M2'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                 top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['M2'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
        ws['M2'].font = Font(name='Century Gothic', size=14, bold=True)
        ws['M2'] = 'Descripcion'

        contador = 3
        for proyectos in ProyectosProyeccionSocial.objects.all():
            # Pintamos los dates en el reporte
            ws.cell(row=contador, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=2).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=2).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=2).value = proyectos.fecha

            ws.cell(row=contador, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=3).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=3).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=3).value = proyectos.programa

            ws.cell(row=contador, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=4).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=4).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=4).value = proyectos.ciudad

            ws.cell(row=contador, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=5).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=5).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=5).value = proyectos.institucion

            ws.cell(row=contador, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=6).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=6).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=6).value = proyectos.numeroEstudiantes

            ws.cell(row=contador, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=7).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=7).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=7).value = proyectos.tipoComunidad

            ws.cell(row=contador, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=8).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=8).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=8).value = proyectos.organizacion

            ws.cell(row=contador, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=9).border = Border(left=Side(border_style="thin"),
                                                            right=Side(border_style="thin"),
                                                            top=Side(border_style="thin"),
                                                            bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=9).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=9).value = proyectos.facultad

            ws.cell(row=contador, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=10).border = Border(left=Side(border_style="thin"),
                                                             right=Side(border_style="thin"),
                                                             top=Side(border_style="thin"),
                                                             bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=10).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=10).value = proyectos.nombreProyecto

            ws.cell(row=contador, column=11).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=11).border = Border(left=Side(border_style="thin"),
                                                             right=Side(border_style="thin"),
                                                             top=Side(border_style="thin"),
                                                             bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=11).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=11).value = proyectos.nombreDocente

            ws.cell(row=contador, column=12).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=12).border = Border(left=Side(border_style="thin"),
                                                             right=Side(border_style="thin"),
                                                             top=Side(border_style="thin"),
                                                             bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=12).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=12).value = proyectos.codigoBput

            ws.cell(row=contador, column=13).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=13).border = Border(left=Side(border_style="thin"),
                                                             right=Side(border_style="thin"),
                                                             top=Side(border_style="thin"),
                                                             bottom=Side(border_style="thin"))
            ws.cell(row=contador, column=13).font = Font(name='Century Gothic', size=14)
            ws.cell(row=contador, column=13).alignment = Alignment(horizontal="center")
            ws.cell(row=contador, column=13).value = proyectos.descripcion

            contador += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "ReporteProyectos.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
